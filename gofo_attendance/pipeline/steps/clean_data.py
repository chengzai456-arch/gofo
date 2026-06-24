"""
步骤1: 数据清洗
输入: 原始数据.xlsx, 离职流程.xlsx, 花名册 (N).xlsx, 0525需剔除GUS白名单人员.xlsx
输出: 清洗后数据.xlsx

清洗步骤（严格按顺序执行）：
  步骤1: 剔除离职人员（最后工作日 < 标准日期；若审批状态有值则额外要求含 审批中/已完成/转交）
  步骤2: 剔除未入职人员（入职日期 > 标准日期）
  步骤3: GL00工号处理（先剔除 GL502563，再仅保留白名单）
  步骤4: 剔除 EU人力资源部
  步骤4.5: 剔除GUS白名单人员（读取白名单文件两个Sheet的工号并剔除）
  步骤5: 补签数匹配
  步骤6: 签字报表累计工时
"""
import os
import pandas as pd
import numpy as np

from ..utils import parse_date

# GL00 白名单
GL00_WHITELIST = {'GL000434', 'GL001344', 'GL000004', 'GL000440', 'GL000446', 'GL000902'}
GL00_EXTRA_REMOVE = {'GL502563'}


def run(config, roster_index=12):
    """
    执行数据清洗

    参数:
        config: 配置字典 (来自 config.get_config())
        roster_index: 花名册文件名序号 (默认 12)
    """
    raw_path    = config['raw_file']
    leave_path  = config['leave_file']
    roster_path = config['roster_pattern'].format(n=roster_index)
    output_file = config['cleaned_file']
    output_dir  = config['workspace']

    # ============================================================
    # 1. 读取文件
    # ============================================================
    df_raw   = pd.read_excel(raw_path, header=0, dtype=str)
    df_leave = pd.read_excel(leave_path, header=0, dtype=str)
    df_roster = pd.read_excel(roster_path, header=0, dtype=str)

    # 去掉第一行重复列头
    if str(df_raw.iloc[0].get('考勤日期', '')).strip() == '考勤日期':
        df_raw = df_raw.iloc[1:].reset_index(drop=True)

    print(f"原始数据行数: {len(df_raw)}")

    # ============================================================
    # 2. 识别标准日期
    # ============================================================
    date_col = '考勤日期'
    df_raw[date_col] = df_raw[date_col].astype(str).str.strip()

    non_empty = df_raw[
        df_raw[date_col].notna() &
        (df_raw[date_col] != '') & (df_raw[date_col] != 'nan') & (df_raw[date_col] != 'NaT')
    ]
    parsed_dates = non_empty[date_col].apply(parse_date).dropna().unique()
    standard_dates = sorted([d for d in parsed_dates if pd.notna(d)])
    standard_date = max(standard_dates) if len(standard_dates) > 1 else standard_dates[0]
    print(f"标准日期: {standard_date.date()}")

    # ============================================================
    # 3. 步骤1: 剔除离职流程
    # ============================================================
    df_leave['_last_work_date'] = df_leave['最后工作日'].apply(parse_date)
    has_status = (
        df_leave['审批状态'].notna().any() and
        df_leave['审批状态'].apply(lambda x: str(x).strip() not in ['', 'nan', 'None']).any()
    )
    if has_status:
        leave_status_filter = ['审批中', '已完成', '转交']
        df_leave['_status_valid'] = df_leave['审批状态'].apply(
            lambda x: any(s in str(x) for s in leave_status_filter) if pd.notna(x) else False
        )
        leave_to_remove = df_leave[
            (df_leave['_last_work_date'] < standard_date) &
            (df_leave['_status_valid'] == True)
        ]['工号'].dropna().unique()
        print(f"步骤1 - 离职剔除(含审批状态过滤): {len(leave_to_remove)}个工号")
    else:
        leave_to_remove = df_leave[
            df_leave['_last_work_date'] < standard_date
        ]['工号'].dropna().unique()
        print(f"步骤1 - 离职剔除(审批状态为空，仅按最后工作日): {len(leave_to_remove)}个工号")
    before = len(df_raw)
    df_raw = df_raw[~df_raw['工号'].isin(leave_to_remove)]
    print(f"  剔除后: {len(df_raw)} (减少 {before - len(df_raw)} 行)")

    # ============================================================
    # 4. 步骤2: 剔除未入职
    # ============================================================
    df_roster['_join_date'] = df_roster['入职日期'].apply(parse_date)
    roster_to_remove = df_roster[
        df_roster['_join_date'] > standard_date
    ]['工号'].dropna().unique()
    print(f"步骤2 - 未入职剔除: {len(roster_to_remove)}个工号")
    before = len(df_raw)
    df_raw = df_raw[~df_raw['工号'].isin(roster_to_remove)]
    print(f"  剔除后: {len(df_raw)} (减少 {before - len(df_raw)} 行)")

    # ============================================================
    # 5. 步骤3: GL00工号处理
    # ============================================================
    before = len(df_raw)
    df_raw = df_raw[~df_raw['工号'].isin(GL00_EXTRA_REMOVE)]
    print(f"步骤3a - 剔除GL502563后: {len(df_raw)} (减少 {before - len(df_raw)} 行)")

    is_gl00 = df_raw['工号'].str.startswith('GL00', na=False)
    not_whitelist = ~df_raw['工号'].isin(GL00_WHITELIST)
    mask_remove = is_gl00 & not_whitelist
    before = len(df_raw)
    df_raw = df_raw[~mask_remove]
    print(f"步骤3b - GL00非白名单剔除: {len(df_raw)} (减少 {before - len(df_raw)} 行)")

    # ============================================================
    # 6. 步骤4: 剔除 EU人力资源部
    # ============================================================
    before = len(df_raw)
    df_raw = df_raw[df_raw['三级部门'].str.strip() != 'EU人力资源部']
    print(f"步骤4 - EU人力资源部剔除: {len(df_raw)} (减少 {before - len(df_raw)} 行)")

    # ============================================================
    # 7. 步骤4.5: 剔除GUS白名单人员
    # ============================================================
    gus_file = config.get('gus_whitelist_file')
    if gus_file and os.path.exists(gus_file):
        before = len(df_raw)
        df_gus = pd.read_excel(gus_file, sheet_name=None, dtype=str)
        # 从所有Sheet收集工号
        gus_ids = set()
        for sheet_name, sheet_df in df_gus.items():
            if '工号' in sheet_df.columns:
                ids = sheet_df['工号'].dropna().unique()
                gus_ids.update(ids)
                print(f"步骤4.5a - Sheet '{sheet_name}' 读取工号数: {len(ids)}")
        print(f"步骤4.5b - GUS白名单去重后: {len(gus_ids)} 个唯一工号")
        df_raw = df_raw[~df_raw['工号'].astype(str).str.strip().isin(gus_ids)]
        print(f"步骤4.5 - GUS白名单剔除: {len(df_raw)} (减少 {before - len(df_raw)} 行)")
    else:
        print(f"步骤4.5 - GUS白名单文件不存在或未配置，跳过")

    # ============================================================
    # 8. 步骤5: 补签数匹配
    # ============================================================
    # 读取补签管理文件，过滤审批状态和日期，按工号计数
    df_resign = pd.read_excel(config['resign_file'], header=0, dtype=str)
    # 兼容新格式：若无审批状态列（透视汇总表），直接跳过补签处理
    if '审批状态' not in df_resign.columns:
        print(f"步骤5 - 补签文件无审批状态列(透视汇总格式), 跳过补签处理")
        df_raw['补签数'] = 0
    else:
        # 审批状态过滤: 仅保留 已完成/审批中/转交
        resign_status_filter = ['已完成', '审批中', '转交']
        df_resign = df_resign[df_resign['审批状态'].apply(
            lambda x: str(x).strip() in resign_status_filter if pd.notna(x) else False
        )]
        print(f"步骤5a - 补签审批状态过滤后: {len(df_resign)} 行 (仅保留 {resign_status_filter})")
        # 补签日期过滤: 仅保留等于标准日期的
        df_resign['_resign_date'] = df_resign['补签日期'].apply(parse_date)
        df_resign = df_resign[df_resign['_resign_date'] == standard_date]
        print(f"步骤5b - 补签日期过滤后(={standard_date.date()}): {len(df_resign)} 行")
        # 按工号计数
        resign_counts = df_resign.groupby('工号').size().reset_index(name='补签数')
        print(f"步骤5c - 补签涉及工号数: {len(resign_counts)}")
        # 合并到清洗后数据: 工号匹配 → 补签数, 无匹配填0
        df_raw = df_raw.merge(resign_counts, on='工号', how='left')
        df_raw['补签数'] = df_raw['补签数'].fillna(0).astype(int)
        print(f"步骤5 - 补签数列已添加 (非零工号数: {(df_raw['补签数'] > 0).sum()})")

    # ============================================================
    # 8. 步骤6: 签字报表累计工时
    # ============================================================
    # 每日总工时列名
    DAILY_HOURS_COL = '每日总工时(公式：末打卡-首打卡-班次午休时间+居家办公时长)'

    def _process_sign_report(file_config_key, output_col, subtract_flag, label):
        """处理签字报表：读取→剔除非蓝色行→按工号汇总每日总工时→计算→合并
        subtract_flag: True=减40取max(0), False=直接累加
        """
        file_path = config.get(file_config_key)
        if not file_path or not os.path.exists(file_path):
            print(f"步骤6-{label} - 跳过(文件不存在): {file_path}")
            return
        # 用 openpyxl 读取以获取单元格颜色，仅保留蓝色行（indexed 填充 = 蓝色合计行）
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        # 找到工号列索引
        emp_col_idx = None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h and str(h).strip() == '工号':
                emp_col_idx = c
                break
        if emp_col_idx is None:
            wb.close()
            df_sign = pd.read_excel(file_path, header=0, dtype=str)
        else:
            blue_emps = set()
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=emp_col_idx)
                fill = cell.fill
                is_indexed = fill and fill.fgColor and fill.fgColor.type == 'indexed'
                emp_val = cell.value
                is_empty = emp_val is None or str(emp_val).strip() == '' or str(emp_val).strip() == 'nan'
                if is_indexed and not is_empty:
                    blue_emps.add(str(emp_val).strip())
            wb.close()
            df_sign = pd.read_excel(file_path, header=0, dtype=str)
            # 用工号匹配而非行号(openpyxl与pandas行号可能不同)
            df_sign = df_sign[df_sign['工号'].isin(blue_emps)].reset_index(drop=True)
            print(f"步骤6-{label}a - 蓝色行过滤: {len(blue_emps)} 工号, {len(df_sign)} 行 (总{ws.max_row - 1}行)")
            # 智能剔除合计行: 按工号分组，若某行的每日总工时 = 其他行之和 → 剔除
            before = len(df_sign)
            df_sign['_tmp_h'] = pd.to_numeric(df_sign[DAILY_HOURS_COL], errors='coerce').fillna(0)
            summary_mask = pd.Series(False, index=df_sign.index)
            for emp, grp in df_sign.groupby('工号'):
                if len(grp) <= 1:
                    continue
                grp_hours = grp['_tmp_h'].values
                for i in grp.index:
                    other_sum = grp_hours[grp.index != i].sum()
                    if abs(grp_hours[grp.index.get_loc(i)] - other_sum) < 0.1:
                        summary_mask[i] = True
            df_sign = df_sign[~summary_mask].drop(columns=['_tmp_h'])
            print(f"步骤6-{label}b - 剔除合计行: {before} → {len(df_sign)} 行 (智能检测)")
        # 汇总每日总工时（蓝色合计行每个员工仅一行，groupby 安全但双周无需聚合）
        df_sign['_hours'] = pd.to_numeric(df_sign[DAILY_HOURS_COL], errors='coerce').fillna(0)
        if subtract_flag:
            # 上周/本周：按工号汇总 → 减40 → floor 0
            emp_sum = df_sign.groupby('工号')['_hours'].sum().reset_index()
            emp_sum[output_col] = (emp_sum['_hours'] - 40).clip(lower=0)
            emp_sum[output_col] = emp_sum[output_col].round(2)
            print(f"步骤6-{label} - 涉及工号数: {len(emp_sum)}, 工时合计: {emp_sum['_hours'].sum():.1f}h")
            merge_df = emp_sum[['工号', output_col]]
        else:
            # 双周：直接使用每日总工时值，按工号去重保留第一条
            emp_sum = df_sign[['工号', '_hours']].copy()
            emp_sum[output_col] = emp_sum['_hours'].astype(float).round(2)
            print(f"步骤6-{label} - 涉及工号数: {len(emp_sum)}, 工时合计: {emp_sum['_hours'].astype(float).sum():.1f}h")
            merge_df = emp_sum[['工号', output_col]].drop_duplicates(subset=['工号'], keep='first')
        # 合并到 df_raw，确保 merge_df 工号唯一
        merge_df = merge_df.drop_duplicates(subset=['工号'], keep='first')
        nonlocal df_raw
        df_raw = df_raw.merge(merge_df, on='工号', how='left')
        df_raw[output_col] = df_raw[output_col].fillna(0)
        print(f"步骤6-{label} - 列已添加 (非零工号数: {(df_raw[output_col] > 0).sum()})")

    _process_sign_report('sign_report_this_week', '本周累计加班工时', True, '本周')
    _process_sign_report('sign_report_last_week', '上周累计加班工时', True, '上周')
    _process_sign_report('sign_report_biweek', '双周累计工时', False, '双周')

    # ============================================================
    # 9. 输出
    # ============================================================
    os.makedirs(output_dir, exist_ok=True)
    df_raw.to_excel(output_file, index=False)
    print(f"\n输出: {output_file}")
    print(f"最终行数: {len(df_raw)}")
