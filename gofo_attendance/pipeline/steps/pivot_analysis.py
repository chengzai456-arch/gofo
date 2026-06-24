"""
步骤3: 透视分析
输入: 指标计算后数据.xlsx
输出: 透视分析.xlsx（含5个Sheet）
"""
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..utils import to_str, to_num


# ============================================================
# 样式
# ============================================================
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_FONT  = Font(color="FFFFFF", bold=True, name="微软雅黑", size=11)
SUMMARY_FILL = PatternFill("solid", fgColor="E2EFDA")
SUMMARY_FONT = Font(bold=True, name="微软雅黑", size=11)
DATA_FONT    = Font(name="微软雅黑", size=11)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
THIN         = Side(style='thin', color='CCCCCC')
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT_FMT      = '0.00%'


def _style_cell(cell, is_summary, num_fmt=None):
    cell.font      = SUMMARY_FONT if is_summary else DATA_FONT
    cell.alignment = CENTER
    cell.border    = BORDER
    if is_summary:
        cell.fill = SUMMARY_FILL
    if num_fmt:
        cell.number_format = num_fmt


def _write_sheet(ws, col_headers, rows_data, rate_cols=None):
    for j, col_name in col_headers:
        c = ws.cell(1, j, col_name)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER

    for i, row in enumerate(rows_data, 2):
        is_sum = row.get('_is_summary', False)
        for j, col_name in col_headers:
            val = row.get(col_name, '')
            c = ws.cell(i, j, val)
            _style_cell(c, is_sum)
        if rate_cols:
            for pct_col, num_key, denom_key in rate_cols:
                num   = float(row.get(num_key, 0) or 0)
                denom = float(row.get(denom_key, 0) or 0)
                rate  = min(num / denom, 1.0) if denom != 0 else None
                c = ws.cell(i, pct_col, rate)
                _style_cell(c, is_sum, PCT_FMT)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 4, 10), 32)
    ws.freeze_panes = 'A2'


def run(config):
    """执行透视分析"""
    input_file  = config['metrics_file']
    output_file = config['pivot_file']

    # ============================================================
    # 1. 读取数据
    # ============================================================
    df = pd.read_excel(input_file, dtype=str)
    print(f"数据行数: {len(df)}")

    # 数值辅助列
    df['_是否日超8H_num'] = (df['是否日超8H'] == '是').astype(int)
    df['_是否排班_num']   = (df['是否排班'] == '是').astype(int)
    df['_正确_num']       = (df['是否排班正确'] == '正确').astype(int)
    df['_不正确_num']     = (df['是否排班正确'] == '不正确').astype(int)
    df['_缺卡数_num']     = df['缺卡数'].apply(to_num)
    df['_标准打卡数_num'] = df['标准打卡数'].apply(to_num)
    df['_打卡数_num']     = df['班次内打卡次数'].apply(to_num)
    df['_日超8H_num']     = df['日超8H'].apply(to_num)
    df['_补签数_num']     = df['补签数'].apply(to_num)

    # ============================================================
    # 2. 构建透视数据
    # ============================================================
    rows_list = []

    for dept3, grp in df.groupby('三级部门'):
        dept3_val = to_str(dept3)
        grp_nonempty = grp[grp['四级部门'].apply(lambda x: to_str(x) != '')]
        grp_empty    = grp[grp['四级部门'].apply(lambda x: to_str(x) == '')]

        # 四级非空：每个四级一条明细行
        for d4, sub in grp_nonempty.groupby('四级部门'):
            sub_unique = sub.drop_duplicates(subset=['工号'])
            cnt    = len(sub_unique)
            over   = int(sub['_是否日超8H_num'].sum())
            over_h = round(sub['_日超8H_num'].sum(), 2)
            shift  = int(sub['_是否排班_num'].sum())
            corr   = int(sub['_正确_num'].sum())
            incorr = int(sub['_不正确_num'].sum())
            miss   = round(sub['_缺卡数_num'].sum(), 2)
            std    = round(sub['_标准打卡数_num'].sum(), 2)
            punch  = round(sub['_打卡数_num'].sum(), 2)  # 打卡数汇总
            # 补签率: 按唯一工号聚合
            emp_agg = sub.groupby('工号').agg(
                buqian=('补签数', 'first'),   # 每个工号的补签数
                queka=('缺卡数', lambda x: x.apply(to_num).sum())  # 每个工号的缺卡数合计
            )
            bq_emp = emp_agg['buqian'].apply(to_num).sum()
            qk_emp = emp_agg['queka'].sum()
            rows_list.append({
                '三级部门': dept3_val, '四级部门': to_str(d4),
                '计数': cnt, '超过人数': over, '超过小时合计': over_h,
                '未超过人数': cnt - over, '已排班数': shift, '未排班数': cnt - shift,
                '正确数量': corr, '不正确数量': incorr, '总计人数': corr + incorr,
                '缺卡数量': miss, '标准打卡数量': std, '打卡数': punch,
                '补签数_emp': bq_emp, '缺卡数_emp': qk_emp,
                '_is_summary': False,
            })

        # 四级为空行: 汇总为 "/"
        if len(grp_empty) > 0:
            cnt    = len(grp_empty.drop_duplicates(subset=['工号']))
            over   = int(grp_empty['_是否日超8H_num'].sum())
            over_h = round(grp_empty['_日超8H_num'].sum(), 2)
            shift  = int(grp_empty['_是否排班_num'].sum())
            corr   = int(grp_empty['_正确_num'].sum())
            incorr = int(grp_empty['_不正确_num'].sum())
            miss   = round(grp_empty['_缺卡数_num'].sum(), 2)
            std    = round(grp_empty['_标准打卡数_num'].sum(), 2)
            punch  = round(grp_empty['_打卡数_num'].sum(), 2)
            emp_agg = grp_empty.groupby('工号').agg(
                buqian=('补签数', 'first'),
                queka=('缺卡数', lambda x: x.apply(to_num).sum())
            )
            bq_emp = emp_agg['buqian'].apply(to_num).sum()
            qk_emp = emp_agg['queka'].sum()
            rows_list.append({
                '三级部门': dept3_val, '四级部门': '/',
                '计数': cnt, '超过人数': over, '超过小时合计': over_h,
                '未超过人数': cnt - over, '已排班数': shift, '未排班数': cnt - shift,
                '正确数量': corr, '不正确数量': incorr, '总计人数': corr + incorr,
                '缺卡数量': miss, '标准打卡数量': std, '打卡数': punch,
                '补签数_emp': bq_emp, '缺卡数_emp': qk_emp,
                '_is_summary': False,
            })

        # 三级部门末尾: 汇总行
        cnt    = len(grp.drop_duplicates(subset=['工号']))
        over   = int(grp['_是否日超8H_num'].sum())
        over_h = round(grp['_日超8H_num'].sum(), 2)
        shift  = int(grp['_是否排班_num'].sum())
        corr   = int(grp['_正确_num'].sum())
        incorr = int(grp['_不正确_num'].sum())
        miss   = round(grp['_缺卡数_num'].sum(), 2)
        std    = round(grp['_标准打卡数_num'].sum(), 2)
        punch  = round(grp['_打卡数_num'].sum(), 2)
        emp_agg = grp.groupby('工号').agg(
            buqian=('补签数', 'first'),
            queka=('缺卡数', lambda x: x.apply(to_num).sum())
        )
        bq_emp = emp_agg['buqian'].apply(to_num).sum()
        qk_emp = emp_agg['queka'].sum()
        rows_list.append({
            '三级部门': dept3_val, '四级部门': '汇总',
            '计数': cnt, '超过人数': over, '超过小时合计': over_h,
            '未超过人数': cnt - over, '已排班数': shift, '未排班数': cnt - shift,
            '正确数量': corr, '不正确数量': incorr, '总计人数': corr + incorr,
            '缺卡数量': miss, '标准打卡数量': std, '打卡数': punch,
            '补签数_emp': bq_emp, '缺卡数_emp': qk_emp,
            '_is_summary': True,
        })

    pivot_df = pd.DataFrame(rows_list)

    # 预计算补签率分母: 补签数(唯一工号) + 缺卡数(唯一工号)
    pivot_df['_补签率分母'] = pivot_df['补签数_emp'] + pivot_df['缺卡数_emp']

    # ============================================================
    # 3. 创建 Excel
    # ============================================================
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("日超8H")
    _write_sheet(ws1, [(1,'三级部门'),(2,'四级部门'),(3,'超过小时合计'),(4,'总计人数')], pivot_df.to_dict('records'))

    ws2 = wb.create_sheet("是否排班")
    _write_sheet(ws2, [(1,'三级部门'),(2,'四级部门'),(3,'已排班数'),(4,'未排班数'),(5,'总计人数'),(6,'排班率')], pivot_df.to_dict('records'), rate_cols=[(6,'已排班数','计数')])

    ws3 = wb.create_sheet("是否排班正确")
    _write_sheet(ws3, [(1,'三级部门'),(2,'四级部门'),(3,'正确数量'),(4,'不正确数量'),(5,'总计人数'),(6,'正确率')], pivot_df.to_dict('records'), rate_cols=[(6,'正确数量','计数')])

    ws4 = wb.create_sheet("打卡率")
    _write_sheet(ws4, [
        (1,'三级部门'),(2,'四级部门'),(3,'打卡数'),(4,'标准打卡数量'),(5,'打卡率'),
        (6,'缺卡数'),(7,'补签数(唯一工号)'),(8,'补签率'),
    ], pivot_df.to_dict('records'), rate_cols=[
        (5,'打卡数','标准打卡数量'),       # 打卡率 = 打卡数 / 标准打卡数
        (8,'补签数_emp','_补签率分母'),     # 补签率 = 补签数(工号汇总) / (补签数+缺卡数)
    ])

    # ============================================================
    # Sheet 5: hub排班是否正确
    # ============================================================
    df_hub = df[df['HUB'].apply(lambda x: to_str(x) == 'hub')].copy()
    hub_rows_list = []

    for dept3, grp in df_hub.groupby('三级部门'):
        dept3_val = to_str(dept3)
        grp_nonempty = grp[grp['五级部门'].apply(lambda x: to_str(x) != '')]
        grp_empty    = grp[grp['五级部门'].apply(lambda x: to_str(x) == '')]

        for d5, sub in grp_nonempty.groupby('五级部门'):
            d4_val = to_str(sub['四级部门'].iloc[0])
            sub_unique = sub.drop_duplicates(subset=['工号'])
            cnt = len(sub_unique)
            corr = int(sub['_正确_num'].sum())
            incorr = int(sub['_不正确_num'].sum())
            hub_rows_list.append({'三级部门': dept3_val, '四级部门': d4_val, '五级部门': to_str(d5), '计数': cnt, '正确': corr, '不正确': incorr, '总计': corr + incorr, '_is_summary': False})

        if len(grp_empty) > 0:
            d4_val = to_str(grp_empty['四级部门'].iloc[0])
            cnt = len(grp_empty.drop_duplicates(subset=['工号']))
            corr = int(grp_empty['_正确_num'].sum())
            incorr = int(grp_empty['_不正确_num'].sum())
            hub_rows_list.append({'三级部门': dept3_val, '四级部门': d4_val, '五级部门': '/', '计数': cnt, '正确': corr, '不正确': incorr, '总计': corr + incorr, '_is_summary': False})

        cnt = len(grp.drop_duplicates(subset=['工号']))
        corr = int(grp['_正确_num'].sum())
        incorr = int(grp['_不正确_num'].sum())
        hub_rows_list.append({'三级部门': dept3_val, '四级部门': '', '五级部门': '汇总', '计数': cnt, '正确': corr, '不正确': incorr, '总计': corr + incorr, '_is_summary': True})

    hub_df = pd.DataFrame(hub_rows_list)
    ws5 = wb.create_sheet("hub排班是否正确")
    _write_sheet(ws5, [(1,'三级部门'),(2,'四级部门'),(3,'五级部门'),(4,'正确'),(5,'不正确'),(6,'总计'),(7,'正确率')], hub_df.to_dict('records'), rate_cols=[(7,'正确','总计')])

    # ============================================================
    # 保存
    # ============================================================
    wb.save(output_file)
    print(f"\n输出: {output_file}")
    total_cnt = int(pivot_df[pivot_df['_is_summary'] == True]['计数'].sum())
    print(f"校验 - 汇总人数合计: {total_cnt} (应={len(df)})")
